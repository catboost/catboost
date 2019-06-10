"""
Module parse to/from Excel
"""

# ---------------------------------------------------------------------
# ExcelFile class
import abc
from datetime import date, datetime, time, timedelta
from distutils.version import LooseVersion
from io import UnsupportedOperation
import os
from textwrap import fill
import warnings

import numpy as np

import pandas._libs.json as json
import pandas.compat as compat
from pandas.compat import (
    OrderedDict, add_metaclass, lrange, map, range, string_types, u, zip)
from pandas.errors import EmptyDataError
from pandas.util._decorators import Appender, deprecate_kwarg

from pandas.core.dtypes.common import (
    is_bool, is_float, is_integer, is_list_like)

from pandas.core import config
from pandas.core.frame import DataFrame

from pandas.io.common import (
    _NA_VALUES, _is_url, _stringify_path, _urlopen, _validate_header_arg,
    get_filepath_or_buffer)
from pandas.io.formats.printing import pprint_thing
from pandas.io.parsers import TextParser

__all__ = ["read_excel", "ExcelWriter", "ExcelFile"]

_writer_extensions = ["xlsx", "xls", "xlsm"]
_writers = {}

_read_excel_doc = """
Read an Excel file into a pandas DataFrame.

Support both `xls` and `xlsx` file extensions from a local filesystem or URL.
Support an option to read a single sheet or a list of sheets.

Parameters
----------
io : str, file descriptor, pathlib.Path, ExcelFile or xlrd.Book
    The string could be a URL. Valid URL schemes include http, ftp, s3,
    gcs, and file. For file URLs, a host is expected. For instance, a local
    file could be /path/to/workbook.xlsx.
sheet_name : str, int, list, or None, default 0
    Strings are used for sheet names. Integers are used in zero-indexed
    sheet positions. Lists of strings/integers are used to request
    multiple sheets. Specify None to get all sheets.

    Available cases:

    * Defaults to ``0``: 1st sheet as a `DataFrame`
    * ``1``: 2nd sheet as a `DataFrame`
    * ``"Sheet1"``: Load sheet with name "Sheet1"
    * ``[0, 1, "Sheet5"]``: Load first, second and sheet named "Sheet5"
      as a dict of `DataFrame`
    * None: All sheets.

header : int, list of int, default 0
    Row (0-indexed) to use for the column labels of the parsed
    DataFrame. If a list of integers is passed those row positions will
    be combined into a ``MultiIndex``. Use None if there is no header.
names : array-like, default None
    List of column names to use. If file contains no header row,
    then you should explicitly pass header=None.
index_col : int, list of int, default None
    Column (0-indexed) to use as the row labels of the DataFrame.
    Pass None if there is no such column.  If a list is passed,
    those columns will be combined into a ``MultiIndex``.  If a
    subset of data is selected with ``usecols``, index_col
    is based on the subset.
parse_cols : int or list, default None
    Alias of `usecols`.

    .. deprecated:: 0.21.0
       Use `usecols` instead.

usecols : int, str, list-like, or callable default None
    Return a subset of the columns.
    * If None, then parse all columns.
    * If int, then indicates last column to be parsed.

    .. deprecated:: 0.24.0
       Pass in a list of int instead from 0 to `usecols` inclusive.

    * If str, then indicates comma separated list of Excel column letters
      and column ranges (e.g. "A:E" or "A,C,E:F"). Ranges are inclusive of
      both sides.
    * If list of int, then indicates list of column numbers to be parsed.
    * If list of string, then indicates list of column names to be parsed.

    .. versionadded:: 0.24.0

    * If callable, then evaluate each column name against it and parse the
      column if the callable returns ``True``.

    .. versionadded:: 0.24.0

squeeze : bool, default False
    If the parsed data only contains one column then return a Series.
dtype : Type name or dict of column -> type, default None
    Data type for data or columns. E.g. {'a': np.float64, 'b': np.int32}
    Use `object` to preserve data as stored in Excel and not interpret dtype.
    If converters are specified, they will be applied INSTEAD
    of dtype conversion.

    .. versionadded:: 0.20.0

engine : str, default None
    If io is not a buffer or path, this must be set to identify io.
    Acceptable values are None or xlrd.
converters : dict, default None
    Dict of functions for converting values in certain columns. Keys can
    either be integers or column labels, values are functions that take one
    input argument, the Excel cell content, and return the transformed
    content.
true_values : list, default None
    Values to consider as True.

    .. versionadded:: 0.19.0

false_values : list, default None
    Values to consider as False.

    .. versionadded:: 0.19.0

skiprows : list-like
    Rows to skip at the beginning (0-indexed).
nrows : int, default None
    Number of rows to parse.

    .. versionadded:: 0.23.0

na_values : scalar, str, list-like, or dict, default None
    Additional strings to recognize as NA/NaN. If dict passed, specific
    per-column NA values. By default the following values are interpreted
    as NaN: '""" + fill("', '".join(sorted(_NA_VALUES)), 70, subsequent_indent="    ") + """'.
keep_default_na : bool, default True
    If na_values are specified and keep_default_na is False the default NaN
    values are overridden, otherwise they're appended to.
verbose : bool, default False
    Indicate number of NA values placed in non-numeric columns.
parse_dates : bool, list-like, or dict, default False
    The behavior is as follows:

    * bool. If True -> try parsing the index.
    * list of int or names. e.g. If [1, 2, 3] -> try parsing columns 1, 2, 3
      each as a separate date column.
    * list of lists. e.g.  If [[1, 3]] -> combine columns 1 and 3 and parse as
      a single date column.
    * dict, e.g. {{'foo' : [1, 3]}} -> parse columns 1, 3 as date and call
      result 'foo'

    If a column or index contains an unparseable date, the entire column or
    index will be returned unaltered as an object data type. For non-standard
    datetime parsing, use ``pd.to_datetime`` after ``pd.read_csv``

    Note: A fast-path exists for iso8601-formatted dates.
date_parser : function, optional
    Function to use for converting a sequence of string columns to an array of
    datetime instances. The default uses ``dateutil.parser.parser`` to do the
    conversion. Pandas will try to call `date_parser` in three different ways,
    advancing to the next if an exception occurs: 1) Pass one or more arrays
    (as defined by `parse_dates`) as arguments; 2) concatenate (row-wise) the
    string values from the columns defined by `parse_dates` into a single array
    and pass that; and 3) call `date_parser` once for each row using one or
    more strings (corresponding to the columns defined by `parse_dates`) as
    arguments.
thousands : str, default None
    Thousands separator for parsing string columns to numeric.  Note that
    this parameter is only necessary for columns stored as TEXT in Excel,
    any numeric columns will automatically be parsed, regardless of display
    format.
comment : str, default None
    Comments out remainder of line. Pass a character or characters to this
    argument to indicate comments in the input file. Any data between the
    comment string and the end of the current line is ignored.
skip_footer : int, default 0
    Alias of `skipfooter`.

    .. deprecated:: 0.23.0
       Use `skipfooter` instead.
skipfooter : int, default 0
    Rows at the end to skip (0-indexed).
convert_float : bool, default True
    Convert integral floats to int (i.e., 1.0 --> 1). If False, all numeric
    data will be read in as floats: Excel stores all numbers as floats
    internally.
mangle_dupe_cols : bool, default True
    Duplicate columns will be specified as 'X', 'X.1', ...'X.N', rather than
    'X'...'X'. Passing in False will cause data to be overwritten if there
    are duplicate names in the columns.
**kwds : optional
        Optional keyword arguments can be passed to ``TextFileReader``.

Returns
-------
DataFrame or dict of DataFrames
    DataFrame from the passed in Excel file. See notes in sheet_name
    argument for more information on when a dict of DataFrames is returned.

See Also
--------
to_excel : Write DataFrame to an Excel file.
to_csv : Write DataFrame to a comma-separated values (csv) file.
read_csv : Read a comma-separated values (csv) file into DataFrame.
read_fwf : Read a table of fixed-width formatted lines into DataFrame.

Examples
--------
The file can be read using the file name as string or an open file object:

>>> pd.read_excel('tmp.xlsx', index_col=0)  # doctest: +SKIP
       Name  Value
0   string1      1
1   string2      2
2  #Comment      3

>>> pd.read_excel(open('tmp.xlsx', 'rb'),
...               sheet_name='Sheet3')  # doctest: +SKIP
   Unnamed: 0      Name  Value
0           0   string1      1
1           1   string2      2
2           2  #Comment      3

Index and header can be specified via the `index_col` and `header` arguments

>>> pd.read_excel('tmp.xlsx', index_col=None, header=None)  # doctest: +SKIP
     0         1      2
0  NaN      Name  Value
1  0.0   string1      1
2  1.0   string2      2
3  2.0  #Comment      3

Column types are inferred but can be explicitly specified

>>> pd.read_excel('tmp.xlsx', index_col=0,
...               dtype={'Name': str, 'Value': float})  # doctest: +SKIP
       Name  Value
0   string1    1.0
1   string2    2.0
2  #Comment    3.0

True, False, and NA values, and thousands separators have defaults,
but can be explicitly specified, too. Supply the values you would like
as strings or lists of strings!

>>> pd.read_excel('tmp.xlsx', index_col=0,
...               na_values=['string1', 'string2'])  # doctest: +SKIP
       Name  Value
0       NaN      1
1       NaN      2
2  #Comment      3

Comment lines in the excel input file can be skipped using the `comment` kwarg

>>> pd.read_excel('tmp.xlsx', index_col=0, comment='#')  # doctest: +SKIP
      Name  Value
0  string1    1.0
1  string2    2.0
2     None    NaN
"""


def register_writer(klass):
    """Adds engine to the excel writer registry. You must use this method to
    integrate with ``to_excel``. Also adds config options for any new
    ``supported_extensions`` defined on the writer."""
    if not compat.callable(klass):
        raise ValueError("Can only register callables as engines")
    engine_name = klass.engine
    _writers[engine_name] = klass
    for ext in klass.supported_extensions:
        if ext.startswith('.'):
            ext = ext[1:]
        if ext not in _writer_extensions:
            config.register_option("io.excel.{ext}.writer".format(ext=ext),
                                   engine_name, validator=str)
            _writer_extensions.append(ext)


def _get_default_writer(ext):
    _default_writers = {'xlsx': 'openpyxl', 'xlsm': 'openpyxl', 'xls': 'xlwt'}
    try:
        import xlsxwriter  # noqa
        _default_writers['xlsx'] = 'xlsxwriter'
    except ImportError:
        pass
    return _default_writers[ext]


def get_writer(engine_name):
    try:
        return _writers[engine_name]
    except KeyError:
        raise ValueError("No Excel writer '{engine}'"
                         .format(engine=engine_name))


@Appender(_read_excel_doc)
@deprecate_kwarg("parse_cols", "usecols")
@deprecate_kwarg("skip_footer", "skipfooter")
def read_excel(io,
               sheet_name=0,
               header=0,
               names=None,
               index_col=None,
               parse_cols=None,
               usecols=None,
               squeeze=False,
               dtype=None,
               engine=None,
               converters=None,
               true_values=None,
               false_values=None,
               skiprows=None,
               nrows=None,
               na_values=None,
               keep_default_na=True,
               verbose=False,
               parse_dates=False,
               date_parser=None,
               thousands=None,
               comment=None,
               skip_footer=0,
               skipfooter=0,
               convert_float=True,
               mangle_dupe_cols=True,
               **kwds):

    # Can't use _deprecate_kwarg since sheetname=None has a special meaning
    if is_integer(sheet_name) and sheet_name == 0 and 'sheetname' in kwds:
        warnings.warn("The `sheetname` keyword is deprecated, use "
                      "`sheet_name` instead", FutureWarning, stacklevel=2)
        sheet_name = kwds.pop("sheetname")

    if 'sheet' in kwds:
        raise TypeError("read_excel() got an unexpected keyword argument "
                        "`sheet`")

    if not isinstance(io, ExcelFile):
        io = ExcelFile(io, engine=engine)

    return io.parse(
        sheet_name=sheet_name,
        header=header,
        names=names,
        index_col=index_col,
        usecols=usecols,
        squeeze=squeeze,
        dtype=dtype,
        converters=converters,
        true_values=true_values,
        false_values=false_values,
        skiprows=skiprows,
        nrows=nrows,
        na_values=na_values,
        keep_default_na=keep_default_na,
        verbose=verbose,
        parse_dates=parse_dates,
        date_parser=date_parser,
        thousands=thousands,
        comment=comment,
        skipfooter=skipfooter,
        convert_float=convert_float,
        mangle_dupe_cols=mangle_dupe_cols,
        **kwds)


class _XlrdReader(object):

    def __init__(self, filepath_or_buffer):
        """Reader using xlrd engine.

        Parameters
        ----------
        filepath_or_buffer : string, path object or Workbook
            Object to be parsed.
        """
        err_msg = "Install xlrd >= 1.0.0 for Excel support"

        try:
            import xlrd
        except ImportError:
            raise ImportError(err_msg)
        else:
            if xlrd.__VERSION__ < LooseVersion("1.0.0"):
                raise ImportError(err_msg +
                                  ". Current version " + xlrd.__VERSION__)

        # If filepath_or_buffer is a url, want to keep the data as bytes so
        # can't pass to get_filepath_or_buffer()
        if _is_url(filepath_or_buffer):
            filepath_or_buffer = _urlopen(filepath_or_buffer)
        elif not isinstance(filepath_or_buffer, (ExcelFile, xlrd.Book)):
            filepath_or_buffer, _, _, _ = get_filepath_or_buffer(
                filepath_or_buffer)

        if isinstance(filepath_or_buffer, xlrd.Book):
            self.book = filepath_or_buffer
        elif not isinstance(filepath_or_buffer, xlrd.Book) and hasattr(
                filepath_or_buffer, "read"):
            # N.B. xlrd.Book has a read attribute too
            if hasattr(filepath_or_buffer, 'seek'):
                try:
                    # GH 19779
                    filepath_or_buffer.seek(0)
                except UnsupportedOperation:
                    # HTTPResponse does not support seek()
                    # GH 20434
                    pass

            data = filepath_or_buffer.read()
            self.book = xlrd.open_workbook(file_contents=data)
        elif isinstance(filepath_or_buffer, compat.string_types):
            self.book = xlrd.open_workbook(filepath_or_buffer)
        else:
            raise ValueError('Must explicitly set engine if not passing in'
                             ' buffer or path for io.')

    @property
    def sheet_names(self):
        return self.book.sheet_names()

    def parse(self,
              sheet_name=0,
              header=0,
              names=None,
              index_col=None,
              usecols=None,
              squeeze=False,
              dtype=None,
              true_values=None,
              false_values=None,
              skiprows=None,
              nrows=None,
              na_values=None,
              verbose=False,
              parse_dates=False,
              date_parser=None,
              thousands=None,
              comment=None,
              skipfooter=0,
              convert_float=True,
              mangle_dupe_cols=True,
              **kwds):

        _validate_header_arg(header)

        from xlrd import (xldate, XL_CELL_DATE,
                          XL_CELL_ERROR, XL_CELL_BOOLEAN,
                          XL_CELL_NUMBER)

        epoch1904 = self.book.datemode

        def _parse_cell(cell_contents, cell_typ):
            """converts the contents of the cell into a pandas
               appropriate object"""

            if cell_typ == XL_CELL_DATE:

                # Use the newer xlrd datetime handling.
                try:
                    cell_contents = xldate.xldate_as_datetime(
                        cell_contents, epoch1904)
                except OverflowError:
                    return cell_contents

                # Excel doesn't distinguish between dates and time,
                # so we treat dates on the epoch as times only.
                # Also, Excel supports 1900 and 1904 epochs.
                year = (cell_contents.timetuple())[0:3]
                if ((not epoch1904 and year == (1899, 12, 31)) or
                        (epoch1904 and year == (1904, 1, 1))):
                    cell_contents = time(cell_contents.hour,
                                         cell_contents.minute,
                                         cell_contents.second,
                                         cell_contents.microsecond)

            elif cell_typ == XL_CELL_ERROR:
                cell_contents = np.nan
            elif cell_typ == XL_CELL_BOOLEAN:
                cell_contents = bool(cell_contents)
            elif convert_float and cell_typ == XL_CELL_NUMBER:
                # GH5394 - Excel 'numbers' are always floats
                # it's a minimal perf hit and less surprising
                val = int(cell_contents)
                if val == cell_contents:
                    cell_contents = val
            return cell_contents

        ret_dict = False

        # Keep sheetname to maintain backwards compatibility.
        if isinstance(sheet_name, list):
            sheets = sheet_name
            ret_dict = True
        elif sheet_name is None:
            sheets = self.book.sheet_names()
            ret_dict = True
        else:
            sheets = [sheet_name]

        # handle same-type duplicates.
        sheets = list(OrderedDict.fromkeys(sheets).keys())

        output = OrderedDict()

        for asheetname in sheets:
            if verbose:
                print("Reading sheet {sheet}".format(sheet=asheetname))

            if isinstance(asheetname, compat.string_types):
                sheet = self.book.sheet_by_name(asheetname)
            else:  # assume an integer if not a string
                sheet = self.book.sheet_by_index(asheetname)

            data = []
            usecols = _maybe_convert_usecols(usecols)

            for i in range(sheet.nrows):
                row = [_parse_cell(value, typ)
                       for value, typ in zip(sheet.row_values(i),
                                             sheet.row_types(i))]
                data.append(row)

            if sheet.nrows == 0:
                output[asheetname] = DataFrame()
                continue

            if is_list_like(header) and len(header) == 1:
                header = header[0]

            # forward fill and pull out names for MultiIndex column
            header_names = None
            if header is not None and is_list_like(header):
                header_names = []
                control_row = [True] * len(data[0])

                for row in header:
                    if is_integer(skiprows):
                        row += skiprows

                    data[row], control_row = _fill_mi_header(data[row],
                                                             control_row)

                    if index_col is not None:
                        header_name, _ = _pop_header_name(data[row], index_col)
                        header_names.append(header_name)

            if is_list_like(index_col):
                # Forward fill values for MultiIndex index.
                if not is_list_like(header):
                    offset = 1 + header
                else:
                    offset = 1 + max(header)

                # Check if we have an empty dataset
                # before trying to collect data.
                if offset < len(data):
                    for col in index_col:
                        last = data[offset][col]

                        for row in range(offset + 1, len(data)):
                            if data[row][col] == '' or data[row][col] is None:
                                data[row][col] = last
                            else:
                                last = data[row][col]

            has_index_names = is_list_like(header) and len(header) > 1

            # GH 12292 : error when read one empty column from excel file
            try:
                parser = TextParser(data,
                                    names=names,
                                    header=header,
                                    index_col=index_col,
                                    has_index_names=has_index_names,
                                    squeeze=squeeze,
                                    dtype=dtype,
                                    true_values=true_values,
                                    false_values=false_values,
                                    skiprows=skiprows,
                                    nrows=nrows,
                                    na_values=na_values,
                                    parse_dates=parse_dates,
                                    date_parser=date_parser,
                                    thousands=thousands,
                                    comment=comment,
                                    skipfooter=skipfooter,
                                    usecols=usecols,
                                    mangle_dupe_cols=mangle_dupe_cols,
                                    **kwds)

                output[asheetname] = parser.read(nrows=nrows)

                if not squeeze or isinstance(output[asheetname], DataFrame):
                    if header_names:
                        output[asheetname].columns = output[
                            asheetname].columns.set_names(header_names)
                    elif compat.PY2:
                        output[asheetname].columns = _maybe_convert_to_string(
                            output[asheetname].columns)

            except EmptyDataError:
                # No Data, return an empty DataFrame
                output[asheetname] = DataFrame()

        if ret_dict:
            return output
        else:
            return output[asheetname]


class ExcelFile(object):
    """
    Class for parsing tabular excel sheets into DataFrame objects.
    Uses xlrd. See read_excel for more documentation

    Parameters
    ----------
    io : string, path object (pathlib.Path or py._path.local.LocalPath),
        file-like object or xlrd workbook
        If a string or path object, expected to be a path to xls or xlsx file.
    engine : string, default None
        If io is not a buffer or path, this must be set to identify io.
        Acceptable values are None or ``xlrd``.
    """

    _engines = {
        'xlrd': _XlrdReader,
    }

    def __init__(self, io, engine=None):
        if engine is None:
            engine = 'xlrd'
        if engine not in self._engines:
            raise ValueError("Unknown engine: {engine}".format(engine=engine))

        # could be a str, ExcelFile, Book, etc.
        self.io = io
        # Always a string
        self._io = _stringify_path(io)

        self._reader = self._engines[engine](self._io)

    def __fspath__(self):
        return self._io

    def parse(self,
              sheet_name=0,
              header=0,
              names=None,
              index_col=None,
              usecols=None,
              squeeze=False,
              converters=None,
              true_values=None,
              false_values=None,
              skiprows=None,
              nrows=None,
              na_values=None,
              parse_dates=False,
              date_parser=None,
              thousands=None,
              comment=None,
              skipfooter=0,
              convert_float=True,
              mangle_dupe_cols=True,
              **kwds):
        """
        Parse specified sheet(s) into a DataFrame

        Equivalent to read_excel(ExcelFile, ...)  See the read_excel
        docstring for more info on accepted parameters
        """

        # Can't use _deprecate_kwarg since sheetname=None has a special meaning
        if is_integer(sheet_name) and sheet_name == 0 and 'sheetname' in kwds:
            warnings.warn("The `sheetname` keyword is deprecated, use "
                          "`sheet_name` instead", FutureWarning, stacklevel=2)
            sheet_name = kwds.pop("sheetname")
        elif 'sheetname' in kwds:
            raise TypeError("Cannot specify both `sheet_name` "
                            "and `sheetname`. Use just `sheet_name`")

        if 'chunksize' in kwds:
            raise NotImplementedError("chunksize keyword of read_excel "
                                      "is not implemented")

        return self._reader.parse(sheet_name=sheet_name,
                                  header=header,
                                  names=names,
                                  index_col=index_col,
                                  usecols=usecols,
                                  squeeze=squeeze,
                                  converters=converters,
                                  true_values=true_values,
                                  false_values=false_values,
                                  skiprows=skiprows,
                                  nrows=nrows,
                                  na_values=na_values,
                                  parse_dates=parse_dates,
                                  date_parser=date_parser,
                                  thousands=thousands,
                                  comment=comment,
                                  skipfooter=skipfooter,
                                  convert_float=convert_float,
                                  mangle_dupe_cols=mangle_dupe_cols,
                                  **kwds)

    @property
    def book(self):
        return self._reader.book

    @property
    def sheet_names(self):
        return self._reader.sheet_names

    def close(self):
        """close io if necessary"""
        if hasattr(self.io, 'close'):
            self.io.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()


def _excel2num(x):
    """
    Convert Excel column name like 'AB' to 0-based column index.

    Parameters
    ----------
    x : str
        The Excel column name to convert to a 0-based column index.

    Returns
    -------
    num : int
        The column index corresponding to the name.

    Raises
    ------
    ValueError
        Part of the Excel column name was invalid.
    """
    index = 0

    for c in x.upper().strip():
        cp = ord(c)

        if cp < ord("A") or cp > ord("Z"):
            raise ValueError("Invalid column name: {x}".format(x=x))

        index = index * 26 + cp - ord("A") + 1

    return index - 1


def _range2cols(areas):
    """
    Convert comma separated list of column names and ranges to indices.

    Parameters
    ----------
    areas : str
        A string containing a sequence of column ranges (or areas).

    Returns
    -------
    cols : list
        A list of 0-based column indices.

    Examples
    --------
    >>> _range2cols('A:E')
    [0, 1, 2, 3, 4]
    >>> _range2cols('A,C,Z:AB')
    [0, 2, 25, 26, 27]
    """
    cols = []

    for rng in areas.split(","):
        if ":" in rng:
            rng = rng.split(":")
            cols.extend(lrange(_excel2num(rng[0]), _excel2num(rng[1]) + 1))
        else:
            cols.append(_excel2num(rng))

    return cols


def _maybe_convert_usecols(usecols):
    """
    Convert `usecols` into a compatible format for parsing in `parsers.py`.

    Parameters
    ----------
    usecols : object
        The use-columns object to potentially convert.

    Returns
    -------
    converted : object
        The compatible format of `usecols`.
    """
    if usecols is None:
        return usecols

    if is_integer(usecols):
        warnings.warn(("Passing in an integer for `usecols` has been "
                       "deprecated. Please pass in a list of int from "
                       "0 to `usecols` inclusive instead."),
                      FutureWarning, stacklevel=2)
        return lrange(usecols + 1)

    if isinstance(usecols, compat.string_types):
        return _range2cols(usecols)

    return usecols


def _validate_freeze_panes(freeze_panes):
    if freeze_panes is not None:
        if (
            len(freeze_panes) == 2 and
            all(isinstance(item, int) for item in freeze_panes)
        ):
            return True

        raise ValueError("freeze_panes must be of form (row, column)"
                         " where row and column are integers")

    # freeze_panes wasn't specified, return False so it won't be applied
    # to output sheet
    return False


def _trim_excel_header(row):
    # trim header row so auto-index inference works
    # xlrd uses '' , openpyxl None
    while len(row) > 0 and (row[0] == '' or row[0] is None):
        row = row[1:]
    return row


def _maybe_convert_to_string(row):
    """
    Convert elements in a row to string from Unicode.

    This is purely a Python 2.x patch and is performed ONLY when all
    elements of the row are string-like.

    Parameters
    ----------
    row : array-like
        The row of data to convert.

    Returns
    -------
    converted : array-like
    """
    if compat.PY2:
        converted = []

        for i in range(len(row)):
            if isinstance(row[i], compat.string_types):
                try:
                    converted.append(str(row[i]))
                except UnicodeEncodeError:
                    break
            else:
                break
        else:
            row = converted

    return row


def _fill_mi_header(row, control_row):
    """Forward fills blank entries in row, but only inside the same parent index

    Used for creating headers in Multiindex.
    Parameters
    ----------
    row : list
        List of items in a single row.
    control_row : list of bool
        Helps to determine if particular column is in same parent index as the
        previous value. Used to stop propagation of empty cells between
        different indexes.

    Returns
    ----------
    Returns changed row and control_row
    """
    last = row[0]
    for i in range(1, len(row)):
        if not control_row[i]:
            last = row[i]

        if row[i] == '' or row[i] is None:
            row[i] = last
        else:
            control_row[i] = False
            last = row[i]

    return _maybe_convert_to_string(row), control_row

# fill blank if index_col not None


def _pop_header_name(row, index_col):
    """
    Pop the header name for MultiIndex parsing.

    Parameters
    ----------
    row : list
        The data row to parse for the header name.
    index_col : int, list
        The index columns for our data. Assumed to be non-null.

    Returns
    -------
    header_name : str
        The extracted header name.
    trimmed_row : list
        The original data row with the header name removed.
    """
    # Pop out header name and fill w/blank.
    i = index_col if not is_list_like(index_col) else max(index_col)

    header_name = row[i]
    header_name = None if header_name == "" else header_name

    return header_name, row[:i] + [''] + row[i + 1:]


@add_metaclass(abc.ABCMeta)
class ExcelWriter(object):
    """
    Class for writing DataFrame objects into excel sheets, default is to use
    xlwt for xls, openpyxl for xlsx.  See DataFrame.to_excel for typical usage.

    Parameters
    ----------
    path : string
        Path to xls or xlsx file.
    engine : string (optional)
        Engine to use for writing. If None, defaults to
        ``io.excel.<extension>.writer``.  NOTE: can only be passed as a keyword
        argument.
    date_format : string, default None
        Format string for dates written into Excel files (e.g. 'YYYY-MM-DD')
    datetime_format : string, default None
        Format string for datetime objects written into Excel files
        (e.g. 'YYYY-MM-DD HH:MM:SS')
    mode : {'w' or 'a'}, default 'w'
        File mode to use (write or append).

    .. versionadded:: 0.24.0

    Attributes
    ----------
    None

    Methods
    -------
    None

    Notes
    -----
    None of the methods and properties are considered public.

    For compatibility with CSV writers, ExcelWriter serializes lists
    and dicts to strings before writing.

    Examples
    --------
    Default usage:

    >>> with ExcelWriter('path_to_file.xlsx') as writer:
    ...     df.to_excel(writer)

    To write to separate sheets in a single file:

    >>> with ExcelWriter('path_to_file.xlsx') as writer:
    ...     df1.to_excel(writer, sheet_name='Sheet1')
    ...     df2.to_excel(writer, sheet_name='Sheet2')

    You can set the date format or datetime format:

    >>> with ExcelWriter('path_to_file.xlsx',
                          date_format='YYYY-MM-DD',
                          datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
    ...     df.to_excel(writer)

    You can also append to an existing Excel file:

    >>> with ExcelWriter('path_to_file.xlsx', mode='a') as writer:
    ...     df.to_excel(writer, sheet_name='Sheet3')
    """
    # Defining an ExcelWriter implementation (see abstract methods for more...)

    # - Mandatory
    #   - ``write_cells(self, cells, sheet_name=None, startrow=0, startcol=0)``
    #     --> called to write additional DataFrames to disk
    #   - ``supported_extensions`` (tuple of supported extensions), used to
    #      check that engine supports the given extension.
    #   - ``engine`` - string that gives the engine name. Necessary to
    #     instantiate class directly and bypass ``ExcelWriterMeta`` engine
    #     lookup.
    #   - ``save(self)`` --> called to save file to disk
    # - Mostly mandatory (i.e. should at least exist)
    #   - book, cur_sheet, path

    # - Optional:
    #   - ``__init__(self, path, engine=None, **kwargs)`` --> always called
    #     with path as first argument.

    # You also need to register the class with ``register_writer()``.
    # Technically, ExcelWriter implementations don't need to subclass
    # ExcelWriter.
    def __new__(cls, path, engine=None, **kwargs):
        # only switch class if generic(ExcelWriter)

        if issubclass(cls, ExcelWriter):
            if engine is None or (isinstance(engine, string_types) and
                                  engine == 'auto'):
                if isinstance(path, string_types):
                    ext = os.path.splitext(path)[-1][1:]
                else:
                    ext = 'xlsx'

                try:
                    engine = config.get_option('io.excel.{ext}.writer'
                                               .format(ext=ext))
                    if engine == 'auto':
                        engine = _get_default_writer(ext)
                except KeyError:
                    error = ValueError("No engine for filetype: '{ext}'"
                                       .format(ext=ext))
                    raise error
            cls = get_writer(engine)

        return object.__new__(cls)

    # declare external properties you can count on
    book = None
    curr_sheet = None
    path = None

    @abc.abstractproperty
    def supported_extensions(self):
        "extensions that writer engine supports"
        pass

    @abc.abstractproperty
    def engine(self):
        "name of engine"
        pass

    @abc.abstractmethod
    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0,
                    freeze_panes=None):
        """
        Write given formatted cells into Excel an excel sheet

        Parameters
        ----------
        cells : generator
            cell of formatted data to save to Excel sheet
        sheet_name : string, default None
            Name of Excel sheet, if None, then use self.cur_sheet
        startrow : upper left cell row to dump data frame
        startcol : upper left cell column to dump data frame
        freeze_panes: integer tuple of length 2
            contains the bottom-most row and right-most column to freeze
        """
        pass

    @abc.abstractmethod
    def save(self):
        """
        Save workbook to disk.
        """
        pass

    def __init__(self, path, engine=None,
                 date_format=None, datetime_format=None, mode='w',
                 **engine_kwargs):
        # validate that this engine can handle the extension
        if isinstance(path, string_types):
            ext = os.path.splitext(path)[-1]
        else:
            ext = 'xls' if engine == 'xlwt' else 'xlsx'

        self.check_extension(ext)

        self.path = path
        self.sheets = {}
        self.cur_sheet = None

        if date_format is None:
            self.date_format = 'YYYY-MM-DD'
        else:
            self.date_format = date_format
        if datetime_format is None:
            self.datetime_format = 'YYYY-MM-DD HH:MM:SS'
        else:
            self.datetime_format = datetime_format

        self.mode = mode

    def __fspath__(self):
        return _stringify_path(self.path)

    def _get_sheet_name(self, sheet_name):
        if sheet_name is None:
            sheet_name = self.cur_sheet
        if sheet_name is None:  # pragma: no cover
            raise ValueError('Must pass explicit sheet_name or set '
                             'cur_sheet property')
        return sheet_name

    def _value_with_fmt(self, val):
        """Convert numpy types to Python types for the Excel writers.

        Parameters
        ----------
        val : object
            Value to be written into cells

        Returns
        -------
        Tuple with the first element being the converted value and the second
            being an optional format
        """
        fmt = None

        if is_integer(val):
            val = int(val)
        elif is_float(val):
            val = float(val)
        elif is_bool(val):
            val = bool(val)
        elif isinstance(val, datetime):
            fmt = self.datetime_format
        elif isinstance(val, date):
            fmt = self.date_format
        elif isinstance(val, timedelta):
            val = val.total_seconds() / float(86400)
            fmt = '0'
        else:
            val = compat.to_str(val)

        return val, fmt

    @classmethod
    def check_extension(cls, ext):
        """checks that path's extension against the Writer's supported
        extensions.  If it isn't supported, raises UnsupportedFiletypeError."""
        if ext.startswith('.'):
            ext = ext[1:]
        if not any(ext in extension for extension in cls.supported_extensions):
            msg = (u("Invalid extension for engine '{engine}': '{ext}'")
                   .format(engine=pprint_thing(cls.engine),
                           ext=pprint_thing(ext)))
            raise ValueError(msg)
        else:
            return True

    # Allow use as a contextmanager
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def close(self):
        """synonym for save, to make it more file-like"""
        return self.save()


class _OpenpyxlWriter(ExcelWriter):
    engine = 'openpyxl'
    supported_extensions = ('.xlsx', '.xlsm')

    def __init__(self, path, engine=None, mode='w', **engine_kwargs):
        # Use the openpyxl module as the Excel writer.
        from openpyxl.workbook import Workbook

        super(_OpenpyxlWriter, self).__init__(path, mode=mode, **engine_kwargs)

        if self.mode == 'a':  # Load from existing workbook
            from openpyxl import load_workbook
            book = load_workbook(self.path)
            self.book = book
        else:
            # Create workbook object with default optimized_write=True.
            self.book = Workbook()

            if self.book.worksheets:
                try:
                    self.book.remove(self.book.worksheets[0])
                except AttributeError:

                    # compat - for openpyxl <= 2.4
                    self.book.remove_sheet(self.book.worksheets[0])

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.save(self.path)

    @classmethod
    def _convert_to_style(cls, style_dict):
        """
        converts a style_dict to an openpyxl style object
        Parameters
        ----------
        style_dict : style dictionary to convert
        """

        from openpyxl.style import Style
        xls_style = Style()
        for key, value in style_dict.items():
            for nk, nv in value.items():
                if key == "borders":
                    (xls_style.borders.__getattribute__(nk)
                     .__setattr__('border_style', nv))
                else:
                    xls_style.__getattribute__(key).__setattr__(nk, nv)

        return xls_style

    @classmethod
    def _convert_to_style_kwargs(cls, style_dict):
        """
        Convert a style_dict to a set of kwargs suitable for initializing
        or updating-on-copy an openpyxl v2 style object
        Parameters
        ----------
        style_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'font'
                'fill'
                'border' ('borders')
                'alignment'
                'number_format'
                'protection'
        Returns
        -------
        style_kwargs : dict
            A dict with the same, normalized keys as ``style_dict`` but each
            value has been replaced with a native openpyxl style object of the
            appropriate class.
        """

        _style_key_map = {
            'borders': 'border',
        }

        style_kwargs = {}
        for k, v in style_dict.items():
            if k in _style_key_map:
                k = _style_key_map[k]
            _conv_to_x = getattr(cls, '_convert_to_{k}'.format(k=k),
                                 lambda x: None)
            new_v = _conv_to_x(v)
            if new_v:
                style_kwargs[k] = new_v

        return style_kwargs

    @classmethod
    def _convert_to_color(cls, color_spec):
        """
        Convert ``color_spec`` to an openpyxl v2 Color object
        Parameters
        ----------
        color_spec : str, dict
            A 32-bit ARGB hex string, or a dict with zero or more of the
            following keys.
                'rgb'
                'indexed'
                'auto'
                'theme'
                'tint'
                'index'
                'type'
        Returns
        -------
        color : openpyxl.styles.Color
        """

        from openpyxl.styles import Color

        if isinstance(color_spec, str):
            return Color(color_spec)
        else:
            return Color(**color_spec)

    @classmethod
    def _convert_to_font(cls, font_dict):
        """
        Convert ``font_dict`` to an openpyxl v2 Font object
        Parameters
        ----------
        font_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'name'
                'size' ('sz')
                'bold' ('b')
                'italic' ('i')
                'underline' ('u')
                'strikethrough' ('strike')
                'color'
                'vertAlign' ('vertalign')
                'charset'
                'scheme'
                'family'
                'outline'
                'shadow'
                'condense'
        Returns
        -------
        font : openpyxl.styles.Font
        """

        from openpyxl.styles import Font

        _font_key_map = {
            'sz': 'size',
            'b': 'bold',
            'i': 'italic',
            'u': 'underline',
            'strike': 'strikethrough',
            'vertalign': 'vertAlign',
        }

        font_kwargs = {}
        for k, v in font_dict.items():
            if k in _font_key_map:
                k = _font_key_map[k]
            if k == 'color':
                v = cls._convert_to_color(v)
            font_kwargs[k] = v

        return Font(**font_kwargs)

    @classmethod
    def _convert_to_stop(cls, stop_seq):
        """
        Convert ``stop_seq`` to a list of openpyxl v2 Color objects,
        suitable for initializing the ``GradientFill`` ``stop`` parameter.
        Parameters
        ----------
        stop_seq : iterable
            An iterable that yields objects suitable for consumption by
            ``_convert_to_color``.
        Returns
        -------
        stop : list of openpyxl.styles.Color
        """

        return map(cls._convert_to_color, stop_seq)

    @classmethod
    def _convert_to_fill(cls, fill_dict):
        """
        Convert ``fill_dict`` to an openpyxl v2 Fill object
        Parameters
        ----------
        fill_dict : dict
            A dict with one or more of the following keys (or their synonyms),
                'fill_type' ('patternType', 'patterntype')
                'start_color' ('fgColor', 'fgcolor')
                'end_color' ('bgColor', 'bgcolor')
            or one or more of the following keys (or their synonyms).
                'type' ('fill_type')
                'degree'
                'left'
                'right'
                'top'
                'bottom'
                'stop'
        Returns
        -------
        fill : openpyxl.styles.Fill
        """

        from openpyxl.styles import PatternFill, GradientFill

        _pattern_fill_key_map = {
            'patternType': 'fill_type',
            'patterntype': 'fill_type',
            'fgColor': 'start_color',
            'fgcolor': 'start_color',
            'bgColor': 'end_color',
            'bgcolor': 'end_color',
        }

        _gradient_fill_key_map = {
            'fill_type': 'type',
        }

        pfill_kwargs = {}
        gfill_kwargs = {}
        for k, v in fill_dict.items():
            pk = gk = None
            if k in _pattern_fill_key_map:
                pk = _pattern_fill_key_map[k]
            if k in _gradient_fill_key_map:
                gk = _gradient_fill_key_map[k]
            if pk in ['start_color', 'end_color']:
                v = cls._convert_to_color(v)
            if gk == 'stop':
                v = cls._convert_to_stop(v)
            if pk:
                pfill_kwargs[pk] = v
            elif gk:
                gfill_kwargs[gk] = v
            else:
                pfill_kwargs[k] = v
                gfill_kwargs[k] = v

        try:
            return PatternFill(**pfill_kwargs)
        except TypeError:
            return GradientFill(**gfill_kwargs)

    @classmethod
    def _convert_to_side(cls, side_spec):
        """
        Convert ``side_spec`` to an openpyxl v2 Side object
        Parameters
        ----------
        side_spec : str, dict
            A string specifying the border style, or a dict with zero or more
            of the following keys (or their synonyms).
                'style' ('border_style')
                'color'
        Returns
        -------
        side : openpyxl.styles.Side
        """

        from openpyxl.styles import Side

        _side_key_map = {
            'border_style': 'style',
        }

        if isinstance(side_spec, str):
            return Side(style=side_spec)

        side_kwargs = {}
        for k, v in side_spec.items():
            if k in _side_key_map:
                k = _side_key_map[k]
            if k == 'color':
                v = cls._convert_to_color(v)
            side_kwargs[k] = v

        return Side(**side_kwargs)

    @classmethod
    def _convert_to_border(cls, border_dict):
        """
        Convert ``border_dict`` to an openpyxl v2 Border object
        Parameters
        ----------
        border_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'left'
                'right'
                'top'
                'bottom'
                'diagonal'
                'diagonal_direction'
                'vertical'
                'horizontal'
                'diagonalUp' ('diagonalup')
                'diagonalDown' ('diagonaldown')
                'outline'
        Returns
        -------
        border : openpyxl.styles.Border
        """

        from openpyxl.styles import Border

        _border_key_map = {
            'diagonalup': 'diagonalUp',
            'diagonaldown': 'diagonalDown',
        }

        border_kwargs = {}
        for k, v in border_dict.items():
            if k in _border_key_map:
                k = _border_key_map[k]
            if k == 'color':
                v = cls._convert_to_color(v)
            if k in ['left', 'right', 'top', 'bottom', 'diagonal']:
                v = cls._convert_to_side(v)
            border_kwargs[k] = v

        return Border(**border_kwargs)

    @classmethod
    def _convert_to_alignment(cls, alignment_dict):
        """
        Convert ``alignment_dict`` to an openpyxl v2 Alignment object
        Parameters
        ----------
        alignment_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'horizontal'
                'vertical'
                'text_rotation'
                'wrap_text'
                'shrink_to_fit'
                'indent'
        Returns
        -------
        alignment : openpyxl.styles.Alignment
        """

        from openpyxl.styles import Alignment

        return Alignment(**alignment_dict)

    @classmethod
    def _convert_to_number_format(cls, number_format_dict):
        """
        Convert ``number_format_dict`` to an openpyxl v2.1.0 number format
        initializer.
        Parameters
        ----------
        number_format_dict : dict
            A dict with zero or more of the following keys.
                'format_code' : str
        Returns
        -------
        number_format : str
        """
        return number_format_dict['format_code']

    @classmethod
    def _convert_to_protection(cls, protection_dict):
        """
        Convert ``protection_dict`` to an openpyxl v2 Protection object.
        Parameters
        ----------
        protection_dict : dict
            A dict with zero or more of the following keys.
                'locked'
                'hidden'
        Returns
        -------
        """

        from openpyxl.styles import Protection

        return Protection(**protection_dict)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0,
                    freeze_panes=None):
        # Write the frame cells using openpyxl.
        sheet_name = self._get_sheet_name(sheet_name)

        _style_cache = {}

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name
            self.sheets[sheet_name] = wks

        if _validate_freeze_panes(freeze_panes):
            wks.freeze_panes = wks.cell(row=freeze_panes[0] + 1,
                                        column=freeze_panes[1] + 1)

        for cell in cells:
            xcell = wks.cell(
                row=startrow + cell.row + 1,
                column=startcol + cell.col + 1
            )
            xcell.value, fmt = self._value_with_fmt(cell.val)
            if fmt:
                xcell.number_format = fmt

            style_kwargs = {}
            if cell.style:
                key = str(cell.style)
                style_kwargs = _style_cache.get(key)
                if style_kwargs is None:
                    style_kwargs = self._convert_to_style_kwargs(cell.style)
                    _style_cache[key] = style_kwargs

            if style_kwargs:
                for k, v in style_kwargs.items():
                    setattr(xcell, k, v)

            if cell.mergestart is not None and cell.mergeend is not None:

                wks.merge_cells(
                    start_row=startrow + cell.row + 1,
                    start_column=startcol + cell.col + 1,
                    end_column=startcol + cell.mergeend + 1,
                    end_row=startrow + cell.mergestart + 1
                )

                # When cells are merged only the top-left cell is preserved
                # The behaviour of the other cells in a merged range is
                # undefined
                if style_kwargs:
                    first_row = startrow + cell.row + 1
                    last_row = startrow + cell.mergestart + 1
                    first_col = startcol + cell.col + 1
                    last_col = startcol + cell.mergeend + 1

                    for row in range(first_row, last_row + 1):
                        for col in range(first_col, last_col + 1):
                            if row == first_row and col == first_col:
                                # Ignore first cell. It is already handled.
                                continue
                            xcell = wks.cell(column=col, row=row)
                            for k, v in style_kwargs.items():
                                setattr(xcell, k, v)


register_writer(_OpenpyxlWriter)


class _XlwtWriter(ExcelWriter):
    engine = 'xlwt'
    supported_extensions = ('.xls',)

    def __init__(self, path, engine=None, encoding=None, mode='w',
                 **engine_kwargs):
        # Use the xlwt module as the Excel writer.
        import xlwt
        engine_kwargs['engine'] = engine

        if mode == 'a':
            raise ValueError('Append mode is not supported with xlwt!')

        super(_XlwtWriter, self).__init__(path, mode=mode, **engine_kwargs)

        if encoding is None:
            encoding = 'ascii'
        self.book = xlwt.Workbook(encoding=encoding)
        self.fm_datetime = xlwt.easyxf(num_format_str=self.datetime_format)
        self.fm_date = xlwt.easyxf(num_format_str=self.date_format)

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.save(self.path)

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0,
                    freeze_panes=None):
        # Write the frame cells using xlwt.

        sheet_name = self._get_sheet_name(sheet_name)

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_sheet(sheet_name)
            self.sheets[sheet_name] = wks

        if _validate_freeze_panes(freeze_panes):
            wks.set_panes_frozen(True)
            wks.set_horz_split_pos(freeze_panes[0])
            wks.set_vert_split_pos(freeze_panes[1])

        style_dict = {}

        for cell in cells:
            val, fmt = self._value_with_fmt(cell.val)

            stylekey = json.dumps(cell.style)
            if fmt:
                stylekey += fmt

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = self._convert_to_style(cell.style, fmt)
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.write_merge(startrow + cell.row,
                                startrow + cell.mergestart,
                                startcol + cell.col,
                                startcol + cell.mergeend,
                                val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)

    @classmethod
    def _style_to_xlwt(cls, item, firstlevel=True, field_sep=',',
                       line_sep=';'):
        """helper which recursively generate an xlwt easy style string
        for example:

            hstyle = {"font": {"bold": True},
            "border": {"top": "thin",
                    "right": "thin",
                    "bottom": "thin",
                    "left": "thin"},
            "align": {"horiz": "center"}}
            will be converted to
            font: bold on; \
                    border: top thin, right thin, bottom thin, left thin; \
                    align: horiz center;
        """
        if hasattr(item, 'items'):
            if firstlevel:
                it = ["{key}: {val}"
                      .format(key=key, val=cls._style_to_xlwt(value, False))
                      for key, value in item.items()]
                out = "{sep} ".format(sep=(line_sep).join(it))
                return out
            else:
                it = ["{key} {val}"
                      .format(key=key, val=cls._style_to_xlwt(value, False))
                      for key, value in item.items()]
                out = "{sep} ".format(sep=(field_sep).join(it))
                return out
        else:
            item = "{item}".format(item=item)
            item = item.replace("True", "on")
            item = item.replace("False", "off")
            return item

    @classmethod
    def _convert_to_style(cls, style_dict, num_format_str=None):
        """
        converts a style_dict to an xlwt style object
        Parameters
        ----------
        style_dict : style dictionary to convert
        num_format_str : optional number format string
        """
        import xlwt

        if style_dict:
            xlwt_stylestr = cls._style_to_xlwt(style_dict)
            style = xlwt.easyxf(xlwt_stylestr, field_sep=',', line_sep=';')
        else:
            style = xlwt.XFStyle()
        if num_format_str is not None:
            style.num_format_str = num_format_str

        return style


register_writer(_XlwtWriter)


class _XlsxStyler(object):
    # Map from openpyxl-oriented styles to flatter xlsxwriter representation
    # Ordering necessary for both determinism and because some are keyed by
    # prefixes of others.
    STYLE_MAPPING = {
        'font': [
            (('name',), 'font_name'),
            (('sz',), 'font_size'),
            (('size',), 'font_size'),
            (('color', 'rgb',), 'font_color'),
            (('color',), 'font_color'),
            (('b',), 'bold'),
            (('bold',), 'bold'),
            (('i',), 'italic'),
            (('italic',), 'italic'),
            (('u',), 'underline'),
            (('underline',), 'underline'),
            (('strike',), 'font_strikeout'),
            (('vertAlign',), 'font_script'),
            (('vertalign',), 'font_script'),
        ],
        'number_format': [
            (('format_code',), 'num_format'),
            ((), 'num_format',),
        ],
        'protection': [
            (('locked',), 'locked'),
            (('hidden',), 'hidden'),
        ],
        'alignment': [
            (('horizontal',), 'align'),
            (('vertical',), 'valign'),
            (('text_rotation',), 'rotation'),
            (('wrap_text',), 'text_wrap'),
            (('indent',), 'indent'),
            (('shrink_to_fit',), 'shrink'),
        ],
        'fill': [
            (('patternType',), 'pattern'),
            (('patterntype',), 'pattern'),
            (('fill_type',), 'pattern'),
            (('start_color', 'rgb',), 'fg_color'),
            (('fgColor', 'rgb',), 'fg_color'),
            (('fgcolor', 'rgb',), 'fg_color'),
            (('start_color',), 'fg_color'),
            (('fgColor',), 'fg_color'),
            (('fgcolor',), 'fg_color'),
            (('end_color', 'rgb',), 'bg_color'),
            (('bgColor', 'rgb',), 'bg_color'),
            (('bgcolor', 'rgb',), 'bg_color'),
            (('end_color',), 'bg_color'),
            (('bgColor',), 'bg_color'),
            (('bgcolor',), 'bg_color'),
        ],
        'border': [
            (('color', 'rgb',), 'border_color'),
            (('color',), 'border_color'),
            (('style',), 'border'),
            (('top', 'color', 'rgb',), 'top_color'),
            (('top', 'color',), 'top_color'),
            (('top', 'style',), 'top'),
            (('top',), 'top'),
            (('right', 'color', 'rgb',), 'right_color'),
            (('right', 'color',), 'right_color'),
            (('right', 'style',), 'right'),
            (('right',), 'right'),
            (('bottom', 'color', 'rgb',), 'bottom_color'),
            (('bottom', 'color',), 'bottom_color'),
            (('bottom', 'style',), 'bottom'),
            (('bottom',), 'bottom'),
            (('left', 'color', 'rgb',), 'left_color'),
            (('left', 'color',), 'left_color'),
            (('left', 'style',), 'left'),
            (('left',), 'left'),
        ],
    }

    @classmethod
    def convert(cls, style_dict, num_format_str=None):
        """
        converts a style_dict to an xlsxwriter format dict

        Parameters
        ----------
        style_dict : style dictionary to convert
        num_format_str : optional number format string
        """

        # Create a XlsxWriter format object.
        props = {}

        if num_format_str is not None:
            props['num_format'] = num_format_str

        if style_dict is None:
            return props

        if 'borders' in style_dict:
            style_dict = style_dict.copy()
            style_dict['border'] = style_dict.pop('borders')

        for style_group_key, style_group in style_dict.items():
            for src, dst in cls.STYLE_MAPPING.get(style_group_key, []):
                # src is a sequence of keys into a nested dict
                # dst is a flat key
                if dst in props:
                    continue
                v = style_group
                for k in src:
                    try:
                        v = v[k]
                    except (KeyError, TypeError):
                        break
                else:
                    props[dst] = v

        if isinstance(props.get('pattern'), string_types):
            # TODO: support other fill patterns
            props['pattern'] = 0 if props['pattern'] == 'none' else 1

        for k in ['border', 'top', 'right', 'bottom', 'left']:
            if isinstance(props.get(k), string_types):
                try:
                    props[k] = ['none', 'thin', 'medium', 'dashed', 'dotted',
                                'thick', 'double', 'hair', 'mediumDashed',
                                'dashDot', 'mediumDashDot', 'dashDotDot',
                                'mediumDashDotDot',
                                'slantDashDot'].index(props[k])
                except ValueError:
                    props[k] = 2

        if isinstance(props.get('font_script'), string_types):
            props['font_script'] = ['baseline', 'superscript',
                                    'subscript'].index(props['font_script'])

        if isinstance(props.get('underline'), string_types):
            props['underline'] = {'none': 0, 'single': 1, 'double': 2,
                                  'singleAccounting': 33,
                                  'doubleAccounting': 34}[props['underline']]

        return props


class _XlsxWriter(ExcelWriter):
    engine = 'xlsxwriter'
    supported_extensions = ('.xlsx',)

    def __init__(self, path, engine=None,
                 date_format=None, datetime_format=None, mode='w',
                 **engine_kwargs):
        # Use the xlsxwriter module as the Excel writer.
        import xlsxwriter

        if mode == 'a':
            raise ValueError('Append mode is not supported with xlsxwriter!')

        super(_XlsxWriter, self).__init__(path, engine=engine,
                                          date_format=date_format,
                                          datetime_format=datetime_format,
                                          mode=mode,
                                          **engine_kwargs)

        self.book = xlsxwriter.Workbook(path, **engine_kwargs)

    def save(self):
        """
        Save workbook to disk.
        """

        return self.book.close()

    def write_cells(self, cells, sheet_name=None, startrow=0, startcol=0,
                    freeze_panes=None):
        # Write the frame cells using xlsxwriter.
        sheet_name = self._get_sheet_name(sheet_name)

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.add_worksheet(sheet_name)
            self.sheets[sheet_name] = wks

        style_dict = {'null': None}

        if _validate_freeze_panes(freeze_panes):
            wks.freeze_panes(*(freeze_panes))

        for cell in cells:
            val, fmt = self._value_with_fmt(cell.val)

            stylekey = json.dumps(cell.style)
            if fmt:
                stylekey += fmt

            if stylekey in style_dict:
                style = style_dict[stylekey]
            else:
                style = self.book.add_format(
                    _XlsxStyler.convert(cell.style, fmt))
                style_dict[stylekey] = style

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.merge_range(startrow + cell.row,
                                startcol + cell.col,
                                startrow + cell.mergestart,
                                startcol + cell.mergeend,
                                cell.val, style)
            else:
                wks.write(startrow + cell.row,
                          startcol + cell.col,
                          val, style)


register_writer(_XlsxWriter)
